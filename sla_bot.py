"""
Основной бот для мониторинга SLA с поддержкой команд
Использует БД для хранения сотрудников, настроек, статусов и шаблонов
"""

import asyncio
import logging
import sys
import os
import re
import io
from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List

from telegram import Bot, Update, ChatMember, InputFile
from telegram.error import TelegramError

# Для Excel отчётов
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import config
from api_client import TaskAPIClient
import db_manager

# Настройка логирования
logging.basicConfig(
    level=getattr(logging, config.LOG_LEVEL),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename=config.LOG_FILE,
    encoding='utf-8'
)
logger = logging.getLogger(__name__)


# ============ ФУНКЦИИ ДЛЯ РАБОТЫ С БД ============

def get_bot_settings() -> Dict[str, Any]:
    """Получить настройки бота из БД"""
    settings = db_manager.get_settings()
    
    return {
        'SLA_HOURS': int(settings.get('SLA_HOURS', 24)),
        'CHECK_INTERVAL_MINUTES': int(settings.get('CHECK_INTERVAL_MINUTES', 180)),
        'TAG_START_HOUR': int(settings.get('TAG_START_HOUR', 9)),
        'TAG_END_HOUR': int(settings.get('TAG_END_HOUR', 18)),
        'TAG_ENABLED': settings.get('TAG_ENABLED', 'True').lower() == 'true',
        'TAG_WORKDAYS_ONLY': settings.get('TAG_WORKDAYS_ONLY', 'True').lower() == 'true',
        'IGNORE_REPLIES': settings.get('IGNORE_REPLIES', 'True').lower() == 'true',
        'IGNORE_EDITS': settings.get('IGNORE_EDITS', 'True').lower() == 'true',
        'IGNORE_FORWARDS': settings.get('IGNORE_FORWARDS', 'True').lower() == 'true',
    }


def get_message_templates() -> Dict[str, str]:
    """Получить все шаблоны сообщений из БД"""
    return db_manager.get_all_templates_dict()


def find_employee_by_name(name_text: str) -> Optional[Dict]:
    """Найти сотрудника по имени (из БД)"""
    return db_manager.get_employee_by_name(name_text)


def find_employees_by_lastname(lastname: str) -> List[Dict]:
    """Найти всех сотрудников по фамилии"""
    employees = db_manager.get_employees(active_only=True)
    result = []
    lastname_lower = lastname.lower().strip()
    
    for emp in employees:
        full_name = emp['full_name'].lower()
        if full_name.startswith(lastname_lower):
            result.append(emp)
        else:
            name_parts = full_name.split()
            if name_parts and name_parts[0] == lastname_lower:
                result.append(emp)
    
    return result


# ============ ОСНОВНОЙ КЛАСС БОТА ============

class SLABot:
    """Бот для мониторинга SLA задач"""
    
    def __init__(self):
        self.bot = Bot(token=config.BOT_TOKEN)
        self.api_client = TaskAPIClient()
        self.chat_id = config.CHAT_ID
        self.notified_tasks = set()
        self.is_running = True
        self.last_update_id = 0
        
        # Инициализируем БД
        db_manager.init_db()
        
        # Загружаем настройки
        self.settings = get_bot_settings()
        self.templates = get_message_templates()
        self.notify_statuses = db_manager.get_notify_statuses()
        
        logger.info(f"✅ Настройки загружены из БД: SLA_HOURS={self.settings['SLA_HOURS']}")
        logger.info(f"📋 Статусы с уведомлениями: {self.notify_statuses}")
    
    def reload_settings(self):
        """Перезагрузить настройки из БД"""
        self.settings = get_bot_settings()
        self.templates = get_message_templates()
        self.notify_statuses = db_manager.get_notify_statuses()
        logger.info("🔄 Настройки перезагружены из БД")
    
    def format_created_date(self, task: Dict) -> str:
        """Форматирует дату создания задачи"""
        created_date = "неизвестно"
        if task.get('created'):
            try:
                created_str = task['created']
                if 'T' in created_str:
                    created_str = created_str.split('+')[0].split('.')[0]
                    created_dt = datetime.strptime(created_str, '%Y-%m-%dT%H:%M:%S')
                    created_date = created_dt.strftime('%d.%m.%Y %H:%M')
            except:
                created_date = str(task['created'])[:16]
        return created_date
    
    def format_reopen_date(self, task: Dict) -> str:
        """Форматирует дату переоткрытия задачи"""
        reopen_date_str = ""
        if task.get('was_reopened') and task.get('reopen_date'):
            try:
                reopen_dt = datetime.fromisoformat(task['reopen_date'].replace('Z', '+00:00'))
                reopen_date_str = reopen_dt.strftime('%d.%m.%Y %H:%M')
            except:
                reopen_date_str = str(task['reopen_date'])[:16]
        return reopen_date_str
    
    async def is_user_admin(self, chat_id: int, user_id: int) -> bool:
        try:
            chat_member = await self.bot.get_chat_member(chat_id, user_id)
            return chat_member.status in [ChatMember.ADMINISTRATOR, ChatMember.OWNER]
        except Exception as e:
            logger.error(f"Ошибка при проверке прав администратора: {e}")
            return False
    
    async def is_allowed_chat(self, chat_id: int) -> bool:
        try:
            chat = await self.bot.get_chat(chat_id)
            if chat.type in ['group', 'supergroup']:
                return True
            else:
                logger.warning(f"Запрещённый чат: {chat_id} (тип: {chat.type})")
                return False
        except Exception as e:
            logger.error(f"Ошибка при проверке типа чата: {e}")
            return False
    
    async def check_tasks(self):
        """Проверяет задачи и отправляет уведомления"""
        if not self.is_running:
            return
        
        self.reload_settings()
        
        logger.info("🔄 Проверка задач...")
        
        try:
            tasks = await self.api_client.get_tasks()
            
            if not tasks:
                logger.info("✅ Нет задач")
                return
            
            # Фильтруем задачи по статусам (только те, где notify_enabled = 1)
            filtered_tasks = []
            for task in tasks:
                task_status = task.get('status', '')
                if task_status in self.notify_statuses:
                    employee = find_employee_by_name(task['assignee'])
                    if employee:
                        # Добавляем created_formatted для каждой задачи
                        task['created_formatted'] = self.format_created_date(task)
                        task['reopen_formatted'] = self.format_reopen_date(task)
                        filtered_tasks.append(task)
                        logger.debug(f"✅ Задача {task['id']} в статусе '{task_status}' — добавлена")
                else:
                    logger.debug(f"⏭️ Задача {task['id']} в статусе '{task_status}' — пропущена")
            
            logger.info(f"📊 Задач в статусах с уведомлениями: {len(filtered_tasks)}")
            
            tasks_to_notify = [t for t in filtered_tasks if t.get('should_notify', False)]
            
            logger.info(f"📊 Задач для уведомления: {len(tasks_to_notify)}")
            
            if not tasks_to_notify:
                logger.info("✅ Нет задач для уведомления")
                return
            
            new_tasks = [t for t in tasks_to_notify if t['id'] not in self.notified_tasks]
            
            if not new_tasks:
                logger.info("✅ Нет новых задач для уведомления")
                return
            
            new_tasks.sort(key=lambda x: x['hours_until_due'])
            
            if len(new_tasks) >= 5:
                logger.info(f"📊 Отправляем Excel отчёт ({len(new_tasks)} задач)")
                await self._send_excel_notification(new_tasks)
            else:
                logger.info(f"📊 Отправляем текстовое уведомление ({len(new_tasks)} задач)")
                await self._send_bulk_notification(new_tasks, is_manual=False)
            
            db_manager.increment_stats('checks')
            
        except Exception as e:
            logger.error(f"❌ Ошибка при проверке задач: {e}", exc_info=True)
    
    async def _send_excel_notification(self, tasks: list):
        """Отправляет Excel файл с тегами исполнителей"""
        if not tasks:
            return
        
        self.reload_settings()
        
        now = datetime.now()
        current_hour = now.hour
        current_weekday = now.weekday()
        
        should_mention = self.settings['TAG_ENABLED']
        if should_mention:
            time_ok = current_hour >= self.settings['TAG_START_HOUR'] and current_hour < self.settings['TAG_END_HOUR']
            day_ok = True
            if self.settings['TAG_WORKDAYS_ONLY']:
                day_ok = current_weekday < 5
            should_mention = time_ok and day_ok
        
        mentions = []
        for task in tasks:
            employee = find_employee_by_name(task['assignee'])
            if employee and should_mention:
                mentions.append(employee['telegram_username'])
        
        mentions_str = " ".join(set(mentions)) if mentions else ""
        
        excel_file = await self._generate_excel_report(tasks)
        
        if mentions_str:
            caption = (
                f"📊 Коллеги, в файле собраны задачи с истекающим SLA ({len(tasks)} шт.).\n"
                f"Просьба обратить внимание на свои задачи: {mentions_str}"
            )
        else:
            caption = (
                f"📊 Коллеги, в файле собраны задачи с истекающим SLA ({len(tasks)} шт.).\n"
                f"Просьба обратить внимание на свои задачи."
            )
        
        try:
            await self.bot.send_document(
                chat_id=self.chat_id,
                document=InputFile(excel_file, filename=excel_file.name),
                caption=caption
            )
            logger.info(f"✅ Отправлен Excel отчёт с {len(tasks)} задачами")
            
            for task in tasks:
                self.notified_tasks.add(task['id'])
                
        except TelegramError as e:
            logger.error(f"❌ Ошибка отправки Excel отчёта: {e}")
    
    async def _send_bulk_notification(self, tasks: list, is_manual: bool = False):
        """Отправляет одно общее уведомление со всеми задачами (текстом)"""
        if not tasks:
            return
        
        self.reload_settings()
        
        now = datetime.now()
        current_hour = now.hour
        current_weekday = now.weekday()
        
        should_mention = self.settings['TAG_ENABLED']
        
        if should_mention:
            time_ok = current_hour >= self.settings['TAG_START_HOUR'] and current_hour < self.settings['TAG_END_HOUR']
            day_ok = True
            if self.settings['TAG_WORKDAYS_ONLY']:
                day_ok = current_weekday < 5
            should_mention = time_ok and day_ok
        
        header = self.templates.get('header', '⚠️ Внимание! Приближается SLA!')
        footer = self.templates.get('footer', 'Коллеги, обратите внимание!')
        task_format = self.templates.get('task_format', '• {title} — исполнитель: {assignee}, дедлайн: {due_date} (осталось {remaining})')
        
        message = f"{header}\n\n"
        messages_sent = 0
        
        for i, task in enumerate(tasks):
            if should_mention:
                employee = find_employee_by_name(task['assignee'])
                if employee:
                    assignee_display = f"{task['assignee']} {employee['telegram_username']}"
                else:
                    assignee_display = task['assignee']
            else:
                assignee_display = task['assignee']
            
            # Получаем статус SLA
            hours = task.get('hours_until_due', 0)
            sla_status_display = self._get_sla_status(hours)
            
            # Информация о переоткрытии
            reopen_info = ""
            if task.get('was_reopened') and task.get('reopen_formatted'):
                reopen_info = f"\n🔄 Переоткрыта: {task.get('reopen_formatted')}"
            
            # Формируем задачу по шаблону
            try:
                task_display = task_format.format(
                    id=task.get('id', ''),
                    title=task.get('title', 'Без названия'),
                    assignee=assignee_display,
                    due_date=task['due_date'].strftime('%d.%m.%Y %H:%M') if task.get('due_date') else 'не указан',
                    remaining=self._format_time(hours),
                    status=task.get('status', 'Неизвестно'),
                    priority=task.get('priority', 'Не указан'),
                    url=task.get('url', ''),
                    created=task.get('created_formatted', 'неизвестно'),
                    sla_status=sla_status_display
                )
            except KeyError as e:
                logger.error(f"❌ Ошибка форматирования: отсутствует переменная {e}")
                # fallback — используем простой формат
                task_display = f"• {task.get('title', 'Без названия')} — {assignee_display}"
            
            # Добавляем информацию о переоткрытии
            if reopen_info:
                task_display += reopen_info
            
            message += task_display + "\n\n"
            
            if i < len(tasks) - 1:
                message += f"{'—' * 45}\n\n"
            
            if not is_manual:
                self.notified_tasks.add(task['id'])
            
            if len(message) > 3500:
                message += f"\n{footer}"
                await self.bot.send_message(
                    chat_id=self.chat_id,
                    text=message,
                    disable_web_page_preview=True
                )
                messages_sent += 1
                logger.info(f"📨 Отправлена часть {messages_sent}")
                await asyncio.sleep(2)
                message = f"{header} (продолжение)\n\n"
        
        if message and not message.endswith(footer):
            message += f"\n{footer}"
        
        if message and len(message) > 0:
            try:
                await self.bot.send_message(
                    chat_id=self.chat_id,
                    text=message,
                    disable_web_page_preview=True
                )
                messages_sent += 1
                logger.info(f"✅ Отправлено общее уведомление с {len(tasks)} задачами")
            except TelegramError as e:
                logger.error(f"❌ Ошибка отправки общего уведомления: {e}")
    
    async def _generate_excel_report(self, tasks: list) -> io.BytesIO:
        """Генерирует Excel файл с отчётом по задачам"""
        logger.info(f"📊 Генерация Excel для {len(tasks)} задач")
        
        if not tasks:
            wb = Workbook()
            ws = wb.active
            ws.title = "SLA Отчёт"
            ws.cell(row=1, column=1, value="Нет задач для отображения")
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            excel_bytes.name = f"sla_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            return excel_bytes
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "SLA Отчёт"
            
            headers = [
                'ID', 'Название', 'Исполнитель', 'Telegram', 'Создана',
                'Дедлайн', 'Ост.(ч)', 'Статус SLA', 'Статус', 'Приоритет', 'Ссылка'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row, task in enumerate(tasks, 2):
                employee = find_employee_by_name(task['assignee'])
                telegram = employee['telegram_username'] if employee else '—'
                hours = task['hours_until_due']
                
                created_date = task.get('created_formatted', '—')
                if created_date == 'неизвестно':
                    created_date = '—'
                
                if hours < 0:
                    sla_status = "ПРОСРОЧЕНО"
                elif hours < 12:
                    sla_status = "Критично"
                elif hours < 24:
                    sla_status = "Скоро"
                else:
                    sla_status = "Норма"
                
                ws.cell(row=row, column=1, value=task['id'])
                ws.cell(row=row, column=2, value=task['title'][:50])
                ws.cell(row=row, column=3, value=task['assignee'])
                ws.cell(row=row, column=4, value=telegram)
                ws.cell(row=row, column=5, value=created_date)
                ws.cell(row=row, column=6, value=task['due_date'].strftime('%d.%m.%Y') if task.get('due_date') else '—')
                ws.cell(row=row, column=7, value=round(hours, 1))
                ws.cell(row=row, column=8, value=sla_status)
                ws.cell(row=row, column=9, value=task['status'][:15])
                ws.cell(row=row, column=10, value=task['priority'] or '—')
                ws.cell(row=row, column=11, value=task['url'])
            
            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 15
            ws.column_dimensions[chr(64 + 11)].width = 30
            
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            excel_bytes.name = f"sla_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            logger.info(f"✅ Excel отчёт сгенерирован")
            return excel_bytes
            
        except Exception as e:
            logger.error(f"❌ Ошибка при генерации Excel: {e}")
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value=f"Ошибка генерации отчёта: {str(e)}")
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            excel_bytes.name = f"sla_report_error.xlsx"
            return excel_bytes
    
    async def _generate_request_excel_report(self, tasks: list, lastname: str = None) -> io.BytesIO:
        """Генерирует Excel файл с отчётом по задачам для команды /request"""
        logger.info(f"📊 Генерация персонального Excel для {len(tasks)} задач")
        
        if not tasks:
            wb = Workbook()
            ws = wb.active
            ws.title = "Задачи"
            ws.cell(row=1, column=1, value="Нет задач для отображения")
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            excel_bytes.name = f"tasks_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            return excel_bytes
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Задачи"
            
            headers = [
                'ID задачи', 'Тип задачи', 'Название', 'Статус',
                'Создана', 'Дедлайн', 'Исполнитель', 'Ссылка'
            ]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row, task in enumerate(tasks, 2):
                issue_type = "Неизвестно"
                if 'raw_data' in task:
                    fields = task['raw_data'].get('fields', {})
                    issue_type_data = fields.get('issuetype', {})
                    issue_type = issue_type_data.get('name', 'Неизвестно')
                
                created_date = "—"
                if 'created' in task and task['created']:
                    try:
                        created_str = task['created']
                        if 'T' in created_str:
                            created_str = created_str.split('+')[0].split('.')[0]
                            created_dt = datetime.strptime(created_str, '%Y-%m-%dT%H:%M:%S')
                            created_date = created_dt.strftime('%d.%m.%Y %H:%M')
                    except:
                        created_date = str(task['created'])[:16]
                
                due_date_str = "—"
                if task.get('due_date'):
                    due_date_str = task['due_date'].strftime('%d.%m.%Y %H:%M')
                
                ws.cell(row=row, column=1, value=task['id'])
                ws.cell(row=row, column=2, value=issue_type)
                ws.cell(row=row, column=3, value=task['title'][:100])
                ws.cell(row=row, column=4, value=task['status'])
                ws.cell(row=row, column=5, value=created_date)
                ws.cell(row=row, column=6, value=due_date_str)
                ws.cell(row=row, column=7, value=task['assignee'])
                ws.cell(row=row, column=8, value=task['url'])
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 40
            
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            
            if lastname:
                safe_name = lastname.replace(' ', '_').replace('@', '')
                excel_bytes.name = f"tasks_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            else:
                excel_bytes.name = f"tasks_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            logger.info(f"✅ Персональный Excel отчёт сгенерирован")
            return excel_bytes
            
        except Exception as e:
            logger.error(f"❌ Ошибка при генерации персонального Excel: {e}")
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value=f"Ошибка генерации отчёта: {str(e)}")
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            excel_bytes.name = f"tasks_error.xlsx"
            return excel_bytes
    
    def _format_time(self, hours: float) -> str:
        if hours < 0:
            return f"⚠️ ПРОСРОЧЕНО на {abs(hours):.1f}ч"
        elif hours < 1:
            minutes = int(hours * 60)
            return f"⏰ {minutes} минут"
        elif hours < 24:
            return f"⏰ {hours:.1f} часов"
        else:
            days = int(hours / 24)
            remaining_hours = hours % 24
            return f"⏰ {days}д {remaining_hours:.0f}ч"
    
    def _get_sla_status(self, hours: float) -> str:
        if hours < 0:
            return "⚠️ ПРОСРОЧЕНО"
        elif hours < 12:
            return "🔴 Критично (менее 12 часов)"
        elif hours < 24:
            return "🟡 Скоро истекает (менее 24 часов)"
        else:
            return "🟢 В норме"
    
    def _format_assignee(self, api_name: str) -> str:
        """Форматирует исполнителя с тегом, если теги включены"""
        if not self.settings.get('TAG_ENABLED', True):
            return api_name
        
        employee = find_employee_by_name(api_name)
        if employee:
            return f"{api_name} {employee['telegram_username']}"
        else:
            return api_name
    
    async def get_task_by_key(self, task_key: str) -> Optional[Dict]:
        try:
            task_data = await self.api_client.get_task_by_key(task_key)
            if not task_data:
                return None
            
            fields = task_data.get('fields', {})
            assignee_data = fields.get('assignee')
            assignee_name = self.api_client._extract_assignee(assignee_data)
            
            due_date, sla_source, remaining_text = self.api_client._extract_sla_date(fields)
            
            now = datetime.now()
            if due_date:
                if due_date.tzinfo is not None:
                    due_date = due_date.replace(tzinfo=None)
                hours_until_due = (due_date - now).total_seconds() / 3600
            else:
                hours_until_due = 9999
            
            was_reopened = False
            reopen_date = None
            try:
                was_reopened, reopen_date = await self.api_client.get_reopen_info(task_key)
            except Exception as e:
                logger.debug(f"Ошибка получения истории для {task_key}: {e}")
            
            created_date = "неизвестно"
            if fields.get('created'):
                try:
                    created_str = fields.get('created')
                    if 'T' in created_str:
                        created_str = created_str.split('+')[0].split('.')[0]
                        created_dt = datetime.strptime(created_str, '%Y-%m-%dT%H:%M:%S')
                        created_date = created_dt.strftime('%d.%m.%Y %H:%M')
                except:
                    created_date = str(fields.get('created'))[:16]
            
            reopen_formatted = ""
            if was_reopened and reopen_date:
                try:
                    reopen_dt = datetime.fromisoformat(reopen_date.replace('Z', '+00:00'))
                    reopen_formatted = reopen_dt.strftime('%d.%m.%Y %H:%M')
                except:
                    reopen_formatted = str(reopen_date)[:16]
            
            task = {
                "id": task_data.get('key'),
                "key": task_data.get('key'),
                "title": fields.get('summary', 'Без названия'),
                "assignee": assignee_name,
                "assignee_raw": assignee_data,
                "due_date": due_date,
                "remaining_text": remaining_text,
                "hours_until_due": hours_until_due,
                "should_notify": False,
                "status": fields.get('status', {}).get('name') if fields.get('status') else 'Неизвестно',
                "status_id": fields.get('status', {}).get('id') if fields.get('status') else None,
                "priority": fields.get('priority', {}).get('name') if fields.get('priority') else None,
                "url": f"{self.api_client.base_url}/browse/{task_data.get('key')}",
                "due_date_source": sla_source,
                "created": fields.get('created'),
                "created_formatted": created_date,
                "was_reopened": was_reopened,
                "reopen_date": reopen_date,
                "reopen_formatted": reopen_formatted,
                "raw_data": task_data
            }
            return task
            
        except Exception as e:
            logger.error(f"Ошибка при получении задачи {task_key}: {e}")
            return None
    
    async def handle_updates(self):
        """Обрабатывает входящие команды"""
        try:
            try:
                updates = await self.bot.get_updates(offset=self.last_update_id + 1, timeout=10)
            except Exception as e:
                logger.debug(f"Ошибка получения обновлений: {e}")
                await asyncio.sleep(1)
                return
            
            for update in updates:
                self.last_update_id = update.update_id
                
                if update.message and update.message.text:
                    text = update.message.text.strip()
                    chat_id = update.message.chat_id
                    user_id = update.message.from_user.id
                    
                    logger.info(f"📨 Получено сообщение: '{text}' от {user_id}")
                    
                    try:
                        is_allowed = await self.is_allowed_chat(chat_id)
                    except Exception as e:
                        logger.debug(f"Ошибка проверки чата: {e}")
                        continue
                    
                    if not is_allowed:
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text="❌ Бот работает только в групповых чатах. Личные сообщения запрещены."
                        )
                        continue
                    
                    # Игнорируем ответы, если настроено
                    if self.settings.get('IGNORE_REPLIES', True) and update.message.reply_to_message:
                        logger.debug("⏭️ Игнорируем ответ на сообщение")
                        continue
                    
                    parts = text.split()
                    full_command = parts[0].lower()
                    
                    if '@' in full_command:
                        base_command = full_command.split('@')[0]
                    else:
                        base_command = full_command
                    
                    if base_command == '/start':
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text=(
                                "✅ Бот мониторинга SLA\n\n"
                                "📋 Доступные команды:\n"
                                "/alarm - показать новые задачи с истекающим SLA\n"
                                "/checking_dep - сформировать Excel отчёт по задачам отдела\n"
                                "/request - выгрузить задачи сотрудника по фамилии (например: /request Бухвиц)\n"
                                "/check - проверить конкретную задачу (Например: /check ZZ-123456)"
                            )
                        )
                    
                    elif base_command == '/help':
                        help_text = (
                            "🤖 Команды бота:\n\n"
                            "/alarm - показать новые задачи с истекающим SLA\n"
                            "/checking_dep - сформировать Excel отчёт по задачам отдела\n"
                            "/request - выгрузить задачи сотрудника по фамилии (например: /request Бухвиц)\n"
                            "/check - проверить конкретную задачу (Например: /check ZZ-12345)"
                        )
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text=help_text
                        )
                    
                    elif base_command == '/alarm':
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text="🔍 Формирую отчёт по новым задачам с истекающим SLA..."
                        )
                        
                        self.reload_settings()
                        tasks = await self.api_client.get_tasks()
                        
                        filtered_tasks = []
                        for task in tasks:
                            task_status = task.get('status', '')
                            if task_status in self.notify_statuses:
                                employee = find_employee_by_name(task['assignee'])
                                if employee:
                                    task['created_formatted'] = self.format_created_date(task)
                                    task['reopen_formatted'] = self.format_reopen_date(task)
                                    filtered_tasks.append(task)
                        
                        tasks_to_notify = [t for t in filtered_tasks if t.get('should_notify', False)]
                        
                        if not tasks_to_notify:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="✅ Нет задач с истекающим SLA"
                            )
                            continue
                        
                        new_tasks = [t for t in tasks_to_notify if t['id'] not in self.notified_tasks]
                        
                        if not new_tasks:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="✅ Нет новых задач с истекающим SLA"
                            )
                            continue
                        
                        new_tasks.sort(key=lambda x: x['hours_until_due'])
                        
                        if len(new_tasks) >= 5:
                            await self._send_excel_notification(new_tasks)
                        else:
                            await self._send_bulk_notification(new_tasks, is_manual=False)
                    
                    elif base_command == '/checking_dep':
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text="📊 Формирую Excel отчёт по задачам отдела..."
                        )
                        
                        self.reload_settings()
                        tasks = await self.api_client.get_tasks()
                        
                        filtered_tasks = []
                        for task in tasks:
                            task_status = task.get('status', '')
                            if task_status in self.notify_statuses:
                                employee = find_employee_by_name(task['assignee'])
                                if employee:
                                    task['created_formatted'] = self.format_created_date(task)
                                    task['reopen_formatted'] = self.format_reopen_date(task)
                                    filtered_tasks.append(task)
                        
                        if not filtered_tasks:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="✅ Нет задач у сотрудников отдела в статусах с уведомлениями"
                            )
                            continue
                        
                        filtered_tasks.sort(key=lambda x: x['hours_until_due'])
                        
                        excel_file = await self._generate_excel_report(filtered_tasks)
                        
                        caption_template = self.templates.get('checking_dep_caption', '📊 Отчёт по задачам отдела (всего: {total})')
                        caption = caption_template.format(total=len(filtered_tasks))
                        
                        await self.bot.send_document(
                            chat_id=chat_id,
                            document=InputFile(excel_file, filename=excel_file.name),
                            caption=caption
                        )
                        
                        logger.info(f"✅ Отправлен Excel отчёт с {len(filtered_tasks)} задачами")
                    
                    elif base_command == '/request':
                        if len(parts) < 2:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="❌ Укажите фамилию сотрудника\n\nПример: /request Бухвиц"
                            )
                            continue
                        
                        lastname = parts[1]
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text=f"🔍 Ищу ВСЕ задачи сотрудников с фамилией '{lastname}'..."
                        )
                        
                        employees_found = find_employees_by_lastname(lastname)
                        
                        if not employees_found:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text=f"❌ Сотрудники с фамилией '{lastname}' не найдены в базе"
                            )
                            continue
                        
                        all_user_tasks = []
                        for emp in employees_found:
                            if 'username' not in emp:
                                logger.warning(f"Для {emp['full_name']} не указан username")
                                continue
                            
                            tasks = await self.api_client.get_all_tasks_by_user(emp['username'])
                            for task in tasks:
                                task_copy = task.copy()
                                task_copy['employee_name'] = emp['full_name']
                                task_copy['created_formatted'] = self.format_created_date(task_copy)
                                task_copy['reopen_formatted'] = self.format_reopen_date(task_copy)
                                all_user_tasks.append(task_copy)
                        
                        if not all_user_tasks:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text=f"✅ У сотрудников с фамилией '{lastname}' нет задач"
                            )
                            continue
                        
                        all_user_tasks.sort(key=lambda x: x.get('created', ''), reverse=True)
                        
                        excel_file = await self._generate_request_excel_report(all_user_tasks, lastname)
                        
                        emp_names = ", ".join([e['full_name'] for e in employees_found if 'username' in e])
                        
                        status_counts = {}
                        for task in all_user_tasks:
                            status = task.get('status', 'Неизвестно')
                            status_counts[status] = status_counts.get(status, 0) + 1
                        
                        status_summary = ", ".join([f"{k}: {v}" for k, v in status_counts.items()])
                        
                        caption_template = self.templates.get('request_caption', '📊 ВСЕ задачи сотрудников: {employees}\n📈 Всего задач: {total}\n📋 {status_summary}')
                        caption = caption_template.format(
                            employees=emp_names,
                            total=len(all_user_tasks),
                            status_summary=status_summary
                        )
                        
                        await self.bot.send_document(
                            chat_id=chat_id,
                            document=InputFile(excel_file, filename=excel_file.name),
                            caption=caption
                        )
                        
                        logger.info(f"✅ Отправлен полный Excel отчёт для фамилии '{lastname}' с {len(all_user_tasks)} задачами")
                    
                    elif base_command == '/check':
                        if len(parts) < 2:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="❌ Укажите номер задачи\n\nПример: /check ZZ-12345"
                            )
                            continue
                        
                        task_key = parts[1].upper()
                        
                        if not re.match(r'^ZZ-\d+$', task_key, re.IGNORECASE):
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="❌ Неверный формат задачи\n\nИспользуйте формат: ZZ-12345"
                            )
                            continue
                        
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text=f"🔍 Ищу задачу {task_key}..."
                        )
                        
                        task = await self.get_task_by_key(task_key)
                        
                        if not task:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text=f"❌ Задача {task_key} не найдена.\n\nВозможно, задача уже закрыта или не существует."
                            )
                            continue
                        
                        # Проверяем, нужно ли показывать теги
                        show_mentions = self.settings.get('TAG_ENABLED', True)
                        
                        if show_mentions:
                            assignee_display = self._format_assignee(task['assignee'])
                        else:
                            assignee_display = task['assignee']
                        
                        # Формируем сообщение по шаблону из БД
                        check_template = self.templates.get('check_task_format', 
                            '📌 Задача: {id}\n📋 Название: {title}\n🔗 Ссылка: {url}\n\n👤 Исполнитель: {assignee}\n📅 Создана: {created}\n⏰ Осталось: {remaining}\n📈 Статус задачи: {status}\n🎯 Приоритет: {priority}')
                        
                        # Определяем статус SLA для отображения
                        sla_status_display = self._get_sla_status(task.get('hours_until_due', 0))
                        
                        task_info = check_template.format(
                            id=task.get('id', ''),
                            title=task.get('title', 'Без названия'),
                            url=task.get('url', ''),
                            assignee=assignee_display,
                            created=task.get('created_formatted', 'неизвестно'),
                            remaining=task.get('remaining_text', self._format_time(task.get('hours_until_due', 0))),
                            status=task.get('status', 'Неизвестно'),
                            priority=task.get('priority', 'Не указан'),
                            sla_status=sla_status_display
                        )
                        
                        # Добавляем информацию о переоткрытии
                        if task.get('was_reopened') and task.get('reopen_formatted'):
                            task_info += f"\n🔄 Переоткрыта: {task.get('reopen_formatted')}"
                        
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text=task_info,
                            disable_web_page_preview=True
                        )
                    
                    elif base_command == '/update':
                        is_admin = await self.is_user_admin(chat_id, user_id)
                        
                        if not is_admin:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="❌ У вас нет прав на выполнение этой команды."
                            )
                            logger.warning(f"Пользователь {user_id} попытался использовать /update без прав админа")
                            continue
                        
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text="🛑 Бот уходит на обновление\n\n⏸ Проверка задач временно приостановлена.\n🔄 Скоро бот будет запущен снова."
                        )
                        logger.info(f"🛑 Останавливаем бота по команде от администратора {user_id}")
                        self.is_running = False
                        await asyncio.sleep(2)
                        sys.exit(0)
                    
                    elif base_command == '/restart':
                        is_admin = await self.is_user_admin(chat_id, user_id)
                        
                        if not is_admin:
                            await self.bot.send_message(
                                chat_id=chat_id,
                                text="❌ У вас нет прав на выполнение этой команды."
                            )
                            logger.warning(f"Пользователь {user_id} попытался использовать /restart без прав админа")
                            continue
                        
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text="🔄 Рестарт бота..."
                        )
                        logger.info(f"🔄 Рестартаем бота по команде от администратора {user_id}")
                        self.is_running = False
                        await asyncio.sleep(2)
                        python = sys.executable
                        os.execl(python, python, *sys.argv)
                    
                    else:
                        await self.bot.send_message(
                            chat_id=chat_id,
                            text="❌ Неизвестная команда\n\nНапишите /help для списка команд"
                        )
                        
        except Exception as e:
            logger.error(f"Ошибка при обработке команд: {e}")
            try:
                await self.bot.send_message(
                    chat_id=chat_id if 'chat_id' in locals() else self.chat_id,
                    text=f"❌ Ошибка при выполнении команды\n\n{str(e)[:200]}"
                )
            except:
                pass
    
    async def run_forever(self):
        """Запускает бесконечный цикл"""
        self.reload_settings()
        check_interval = self.settings['CHECK_INTERVAL_MINUTES']
        
        logger.info(f"🚀 Бот запущен. Интервал проверки: {check_interval} минут")
        logger.info(f"📋 SLA_HOURS: {self.settings['SLA_HOURS']} часов")
        logger.info(f"🔔 Теги включены: {self.settings['TAG_ENABLED']}")
        logger.info(f"📋 Статусы с уведомлениями: {self.notify_statuses}")
        
        self.last_update_id = 0
        
        while self.is_running:
            try:
                current_minute = datetime.now().minute
                if current_minute % check_interval == 0:
                    await self.check_tasks()
                    await asyncio.sleep(60)
                
                await self.handle_updates()
                await asyncio.sleep(1)
                
            except KeyboardInterrupt:
                logger.info("🛑 Бот остановлен")
                break
            except Exception as e:
                logger.error(f"❌ Ошибка в основном цикле: {e}", exc_info=True)
                await asyncio.sleep(5)


async def test_bot():
    """Тестовая функция"""
    print("\n" + "=" * 60)
    print("🤖 ТЕСТИРОВАНИЕ SLA БОТА")
    print("=" * 60)
    
    db_manager.init_db()
    
    settings = get_bot_settings()
    templates = get_message_templates()
    notify_statuses = db_manager.get_notify_statuses()
    
    print("\n📋 Настройки из БД:")
    print(f"   SLA_HOURS: {settings['SLA_HOURS']}")
    print(f"   CHECK_INTERVAL_MINUTES: {settings['CHECK_INTERVAL_MINUTES']}")
    print(f"   TAG_ENABLED: {settings['TAG_ENABLED']}")
    
    print(f"\n📋 Статусы с уведомлениями: {notify_statuses}")
    
    print(f"\n📝 Шаблоны в БД:")
    for name, template in templates.items():
        print(f"   {name}: {template[:50]}...")
    
    employees = db_manager.get_employees(active_only=True)
    print(f"\n👥 Сотрудников в БД: {len(employees)}")
    
    print("\n🔍 Получаем задачи из Jira...")
    bot = SLABot()
    tasks = await bot.api_client.get_tasks()
    
    if tasks:
        print(f"\n✅ Получено задач: {len(tasks)}")
        
        filtered = []
        for task in tasks:
            if task.get('status', '') in notify_statuses:
                filtered.append(task)
        
        print(f"📊 После фильтрации по статусам: {len(filtered)}")
        
        to_notify = [t for t in filtered if t.get('should_notify')]
        print(f"⚠️ Требуют уведомления: {len(to_notify)}")
        
        if tasks:
            print(f"\n📋 Пример задачи:")
            task = tasks[0]
            print(f"   ID: {task['id']}")
            print(f"   Статус: {task['status']}")
            print(f"   Исполнитель: {task['assignee']}")
            print(f"   Дедлайн: {task['due_date'].strftime('%d.%m.%Y %H:%M') if task.get('due_date') else 'Нет'}")
    else:
        print("\n❌ Не удалось получить задачи")
    
    print("\n" + "=" * 60)


async def send_test_notification():
    """Отправляет тестовое уведомление"""
    print("\n📨 ОТПРАВКА ТЕСТОВОГО УВЕДОМЛЕНИЯ")
    
    bot = SLABot()
    
    employee = find_employee_by_name("Бухвиц Владислав")
    if not employee:
        print("❌ Сотрудник не найден в БД")
        return
    
    test_task = {
        "id": "TEST-001",
        "title": "🔧 ТЕСТОВАЯ ЗАДАЧА",
        "assignee": "Бухвиц Владислав",
        "due_date": datetime.now() + timedelta(hours=2),
        "hours_until_due": 2.5,
        "status": "В процессе",
        "priority": "High",
        "url": "https://test.ru",
        "created": datetime.now().isoformat(),
        "created_formatted": datetime.now().strftime('%d.%m.%Y %H:%M'),
        "was_reopened": True,
        "reopen_formatted": (datetime.now() - timedelta(days=1)).strftime('%d.%m.%Y %H:%M')
    }
    
    await bot._send_bulk_notification([test_task])
    print("✅ Тестовое уведомление отправлено!")


if __name__ == "__main__":
    db_manager.init_db()
    
    if len(sys.argv) > 1:
        if sys.argv[1] == "--test":
            asyncio.run(test_bot())
        elif sys.argv[1] == "--send-test":
            asyncio.run(send_test_notification())
    else:
        bot = SLABot()
        try:
            asyncio.run(bot.run_forever())
        except KeyboardInterrupt:
            print("\n🛑 Бот остановлен")
