"""
Веб-интерфейс для управления SLA ботом
"""

from flask import Flask, render_template, request, jsonify
from datetime import datetime
import os
import time
import logging
import re

import db_manager

app = Flask(__name__)

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ============ ФУНКЦИЯ ПАРСИНГА ОШИБОК ============

def parse_and_save_errors():
    """Парсит лог и сохраняет ошибки в БД (если их там ещё нет)"""
    log_file = 'sla_bot.log'
    if not os.path.exists(log_file):
        return
    
    error_map = {
        "KeyError": "В шаблоне используется переменная, которой нет в данных. Уберите её из шаблона в админке.",
        "ConnectionError": "Проверьте доступность Jira и настройки подключения в config.py",
        "TelegramError": "Проверьте CHAT_ID и BOT_TOKEN в config.py.",
        "sqlite3.OperationalError": "Ошибка в структуре БД. Удалите settings.db и перезапустите бота.",
        "ModuleNotFoundError": "Не установлен модуль. Установите: pip install <module>",
        "TimeoutError": "Таймаут подключения. Проверьте интернет-соединение.",
        "JSONDecodeError": "Jira вернул невалидный JSON. Проверьте ответ Jira.",
        "PermissionError": "Нет прав на запись в файл. Проверьте права на папку."
    }
    
    try:
        with open(log_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            last_lines = lines[-500:] if len(lines) > 500 else lines
            
            for line in last_lines:
                if 'ERROR' in line or 'Exception' in line or 'Traceback' in line:
                    existing = db_manager.get_error_logs('all')
                    exists = any(line.strip() == e['message'] for e in existing)
                    if exists:
                        continue
                    
                    timestamp = line[:19] if len(line) > 19 else ''
                    solution = 'Обратитесь к администратору'
                    
                    for key, sol in error_map.items():
                        if key in line:
                            solution = sol
                            break
                    
                    db_manager.save_error_log(timestamp, line.strip(), solution)
    except Exception as e:
        logger.error(f"Ошибка парсинга логов: {e}")


# ============ ИНИЦИАЛИЗАЦИЯ ============

db_manager.init_db()
parse_and_save_errors()

# Переносим сотрудников из employees.py если они есть
try:
    from employees import EMPLOYEES
    existing = db_manager.get_employees(active_only=False)
    if len(existing) == 0:
        for emp in EMPLOYEES:
            db_manager.add_employee({
                'full_name': emp['full_name'],
                'search_names': emp['search_names'],
                'telegram_username': emp['telegram_username'],
                'email': emp['email'],
                'username': emp.get('username', emp['email'].split('@')[0]),
                'status': 'active'
            })
        print("✅ Сотрудники перенесены из employees.py")
except ImportError:
    pass

# ============ СТРАНИЦЫ ============

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login')
def login():
    return render_template('index.html')


# ============ API: СОТРУДНИКИ ============

@app.route('/api/employees')
def api_get_employees():
    employees = db_manager.get_employees(active_only=True)
    return jsonify(employees)

@app.route('/api/employees/<int:employee_id>')
def api_get_employee(employee_id):
    employee = db_manager.get_employee_by_id(employee_id)
    if employee:
        return jsonify(employee)
    return jsonify({'error': 'Сотрудник не найден'}), 404

@app.route('/api/employees', methods=['POST'])
def api_add_employee():
    data = request.json
    
    if 'username' not in data or not data['username']:
        data['username'] = data['email'].split('@')[0] if data.get('email') else ''
    
    if 'search_names' not in data or not data['search_names']:
        name_parts = data['full_name'].split()
        data['search_names'] = [name_parts[0], name_parts[1]] if len(name_parts) >= 2 else [name_parts[0]]
    
    if 'status' not in data:
        data['status'] = 'active'
    
    # Добавляем vacation_end_date если есть
    if 'vacation_end_date' not in data or not data['vacation_end_date']:
        data['vacation_end_date'] = None
    
    db_manager.add_employee(data)
    return jsonify({'status': 'ok', 'message': 'Сотрудник добавлен'})

@app.route('/api/employees/<int:employee_id>', methods=['PUT'])
def api_update_employee(employee_id):
    data = request.json
    
    if 'search_names' not in data or not data['search_names']:
        name_parts = data['full_name'].split()
        data['search_names'] = [name_parts[0], name_parts[1]] if len(name_parts) >= 2 else [name_parts[0]]
    
    if 'status' not in data:
        data['status'] = 'active'
    
    if 'vacation_end_date' not in data or not data['vacation_end_date']:
        data['vacation_end_date'] = None
    
    db_manager.update_employee(employee_id, data)
    return jsonify({'status': 'ok', 'message': 'Сотрудник обновлён'})

@app.route('/api/employees/<int:employee_id>', methods=['DELETE'])
def api_delete_employee(employee_id):
    db_manager.delete_employee(employee_id, soft=True)
    return jsonify({'status': 'ok', 'message': 'Сотрудник удалён'})


# ============ API: НАСТРОЙКИ ============

@app.route('/api/settings')
def api_get_settings():
    settings = db_manager.get_settings()
    return jsonify(settings)

@app.route('/api/settings', methods=['POST'])
def api_update_settings():
    data = request.json
    db_manager.update_settings(data)
    return jsonify({'status': 'ok', 'message': 'Настройки сохранены'})


# ============ API: СТАТУСЫ ============

@app.route('/api/statuses')
def api_get_statuses():
    statuses = db_manager.get_task_statuses(active_only=True)
    return jsonify(statuses)

@app.route('/api/statuses', methods=['POST'])
def api_add_status():
    data = request.json
    db_manager.add_task_status(data['name'], data.get('notify_enabled', 1))
    return jsonify({'status': 'ok', 'message': 'Статус добавлен'})

@app.route('/api/statuses/<name>', methods=['PUT'])
def api_update_status(name):
    data = request.json
    db_manager.update_task_status(name, data.get('notify_enabled', 1))
    return jsonify({'status': 'ok', 'message': 'Статус обновлён'})

@app.route('/api/statuses/<name>', methods=['DELETE'])
def api_delete_status(name):
    db_manager.delete_task_status(name, soft=True)
    return jsonify({'status': 'ok', 'message': 'Статус удалён'})


# ============ API: ШАБЛОНЫ ============

@app.route('/api/templates')
def api_get_templates():
    templates = db_manager.get_all_templates_dict()
    return jsonify(templates)

@app.route('/api/templates', methods=['POST'])
def api_save_templates():
    data = request.json
    db_manager.save_templates(data)
    return jsonify({'status': 'ok', 'message': 'Шаблоны сохранены'})


# ============ API: ЛОГИ ============

@app.route('/api/logs')
def api_get_logs():
    """Возвращает логи (последние 500 строк)"""
    log_file = 'sla_bot.log'
    if not os.path.exists(log_file):
        return 'Лог-файл не найден'
    
    try:
        with open(log_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            last_lines = lines[-500:] if len(lines) > 500 else lines
            return ''.join(last_lines)
    except Exception as e:
        return f'Ошибка чтения логов: {e}'

@app.route('/api/logs', methods=['DELETE'])
def api_clear_logs():
    log_file = 'sla_bot.log'
    if os.path.exists(log_file):
        with open(log_file, 'w', encoding='utf-8') as f:
            f.write('')
    return jsonify({'status': 'ok', 'message': 'Логи очищены'})


# ============ API: ОШИБКИ ============

@app.route('/api/error-logs')
def api_error_logs():
    """Возвращает логи ошибок с расшифровкой и статусами"""
    status_filter = request.args.get('status', 'active')
    errors = db_manager.get_error_logs(status_filter)
    return jsonify(errors)


@app.route('/api/error-logs/<int:error_id>/status', methods=['PUT'])
def api_update_error_status(error_id):
    """Обновить статус ошибки"""
    data = request.json
    new_status = data.get('status')
    
    if new_status not in ['new', 'in_progress', 'done']:
        return jsonify({'error': 'Неверный статус'}), 400
    
    db_manager.update_error_status(error_id, new_status)
    return jsonify({'status': 'ok', 'message': 'Статус обновлён'})


@app.route('/api/error-logs/clear-done', methods=['DELETE'])
def api_clear_done_errors():
    """Удалить выполненные ошибки"""
    deleted = db_manager.delete_done_errors()
    return jsonify({'status': 'ok', 'message': f'Удалено {deleted} выполненных ошибок'})


# ============ API: СТАТИСТИКА ============

@app.route('/api/stats')
def api_get_stats():
    """Возвращает статистику (без обращения к Jira)"""
    stats = db_manager.get_stats()
    
    employees = db_manager.get_employees(active_only=True)
    stats['employees_count'] = len(employees)
    
    try:
        result = os.popen('pgrep -f "sla_bot.py"').read().strip()
        stats['bot_running'] = bool(result)
        stats['bot_pid'] = result if result else None
    except:
        stats['bot_running'] = False
        stats['bot_pid'] = None
    
    stats['total_tasks'] = 0
    stats['urgent_tasks'] = 0
    
    errors = db_manager.get_error_logs('active')
    stats['active_errors'] = len(errors)
    
    try:
        stats['history_count'] = db_manager.get_notification_stats()['total']
        stats['pending_count'] = db_manager.get_notification_stats()['pending']
    except:
        stats['history_count'] = 0
        stats['pending_count'] = 0
    
    return jsonify(stats)


# ============ API: УПРАВЛЕНИЕ БОТОМ ============

@app.route('/api/restart', methods=['POST'])
def api_restart_bot():
    try:
        logger.info("🔄 Перезапуск бота...")
        
        os.system('pkill -f "sla_bot.py"')
        time.sleep(2)
        
        command = 'cd /root/sla-monitor-bot && source venv/bin/activate && nohup python3 sla_bot.py >> bot.log 2>&1 &'
        os.system(f'bash -c "{command}"')
        
        time.sleep(2)
        
        result = os.popen('pgrep -f "sla_bot.py"').read().strip()
        if result:
            logger.info(f"✅ Бот перезапущен (PID: {result})")
            return jsonify({'status': 'ok', 'message': f'🔄 Бот перезапущен (PID: {result})'})
        else:
            logger.error("❌ Бот не запустился")
            return jsonify({'status': 'error', 'message': '❌ Бот не запустился'}), 500
            
    except Exception as e:
        logger.error(f"❌ Ошибка перезапуска: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/stop', methods=['POST'])
def api_stop_bot():
    try:
        os.system('pkill -f "sla_bot.py"')
        logger.info("⏹ Бот остановлен")
        return jsonify({'status': 'ok', 'message': '⏹ Бот остановлен'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============ API: ПИНГ ============

@app.route('/api/bot-ping')
def api_bot_ping():
    """Упрощённый пинг — только проверка процесса"""
    import os
    result = os.popen('pgrep -f "sla_bot.py"').read().strip()
    if result:
        return jsonify({
            'status': 'ok',
            'message': '✅ OK',
            'pid': result,
            'timestamp': datetime.now().isoformat()
        })
    else:
        return jsonify({
            'status': 'error',
            'message': '❌ Бот не запущен'
        }), 503


# ============ API: УВЕДОМЛЕНИЯ ============

@app.route('/api/task/<task_key>')
def api_get_task(task_key):
    """Получить информацию о задаче по ключу"""
    try:
        import asyncio
        from api_client import TaskAPIClient
        
        async def get_task():
            client = TaskAPIClient()
            return await client.get_task_by_key(task_key)
        
        task = asyncio.run(get_task())
        
        if task:
            if task.get('due_date'):
                task['due_date'] = task['due_date'].isoformat() if hasattr(task['due_date'], 'isoformat') else str(task['due_date'])
            
            if task.get('created'):
                try:
                    from datetime import datetime
                    created = datetime.fromisoformat(task['created'].replace('Z', '+00:00'))
                    task['created_formatted'] = created.strftime('%d.%m.%Y %H:%M')
                except:
                    task['created_formatted'] = str(task['created'])[:16]
            
            return jsonify(task)
        else:
            return jsonify({'error': 'Задача не найдена'}), 404
            
    except Exception as e:
        logger.error(f"❌ Ошибка получения задачи: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/send-notification', methods=['POST'])
def api_send_notification():
    """Отправить уведомление по задаче"""
    try:
        import asyncio
        import requests
        import config
        from api_client import TaskAPIClient
        import db_manager
        
        data = request.json
        task_key = data.get('task_key')
        priority = data.get('priority', 'Обычное')
        
        if not task_key:
            return jsonify({'error': 'Не указан номер задачи'}), 400
        
        async def get_task():
            client = TaskAPIClient()
            return await client.get_task_by_key(task_key)
        
        task = asyncio.run(get_task())
        
        if not task:
            return jsonify({'error': 'Задача не найдена'}), 404
        
        employee = db_manager.get_employee_by_name(task['assignee'])
        
        if employee and employee.get('status') == 'active':
            mention = employee['telegram_username']
        else:
            mention = task['assignee']
            logger.info(f"⏭️ Тег пропущен: сотрудник {task['assignee']} в статусе '{employee.get('status') if employee else 'не найден'}'")
        
        priority_emoji = {
            'Обычное': '📨',
            'Важное': '⚠️',
            'Срочное': '🚨',
            'Критичное': '🔥'
        }
        
        priority_header = {
            'Обычное': 'Уведомление по задаче',
            'Важное': 'ВАЖНОЕ УВЕДОМЛЕНИЕ!',
            'Срочное': 'СРОЧНОЕ УВЕДОМЛЕНИЕ! 🚨',
            'Критичное': 'КРИТИЧНОЕ УВЕДОМЛЕНИЕ! 🔥'
        }
        
        emoji = priority_emoji.get(priority, '📨')
        header = priority_header.get(priority, 'Уведомление по задаче')
        
        due_date_str = task['due_date'].strftime('%d.%m.%Y %H:%M') if task.get('due_date') else 'не указан'
        
        message = f"{emoji} {header}\n\n"
        message += f"📌 Задача: {task['id']}\n"
        message += f"🔗 Ссылка: {task['url']}\n"
        message += f"📋 Название: {task['title']}\n"
        message += f"👤 Исполнитель: {task['assignee']} {mention if mention != task['assignee'] else ''}\n"
        message += f"📈 Статус: {task['status']}\n"
        message += f"⏰ Дедлайн: {due_date_str}\n"
        message += f"🎯 Приоритет: {task['priority'] or 'Не указан'}\n\n"
        
        if priority == 'Важное':
            message += "❗ Просьба обратить внимание на задачу!"
        elif priority == 'Срочное':
            message += "❗ Требуется срочное внимание к задаче!"
        elif priority == 'Критичное':
            message += "⛔ Задача требует немедленного решения!"
        else:
            message += "Просьба обратить внимание на задачу."
        
        bot_token = config.BOT_TOKEN
        chat_id = config.CHAT_ID
        
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": message,
            "disable_web_page_preview": True
        }
        
        response = requests.post(url, json=payload, timeout=10)
        
        if response.status_code == 200:
            db_manager.save_notification_history(
                message_text=message,
                tasks=[task],
                is_excel=False,
                status='manual'
            )
            return jsonify({'status': 'ok', 'message': '✅ Уведомление отправлено'})
        else:
            logger.error(f"❌ Ошибка Telegram API: {response.status_code} - {response.text}")
            return jsonify({'error': f'Ошибка Telegram API: {response.status_code}'}), 500
        
    except Exception as e:
        logger.error(f"❌ Ошибка отправки уведомления: {e}")
        return jsonify({'error': str(e)}), 500


# ============ API: ИСТОРИЯ УВЕДОМЛЕНИЙ ============

@app.route('/api/notification-history')
def api_notification_history():
    """Получить историю уведомлений с фильтром по статусу"""
    limit = request.args.get('limit', 100, type=int)
    search = request.args.get('search', '')
    status_filter = request.args.get('status', 'all')
    history = db_manager.get_notification_history(limit=limit, search=search, status_filter=status_filter)
    return jsonify(history)


@app.route('/api/notification-history/<int:history_id>/resend', methods=['POST'])
def api_resend_notification(history_id):
    """Повторно отправить уведомление"""
    try:
        import requests
        import config
        
        record = db_manager.get_notification_by_id(history_id)
        if not record:
            return jsonify({'error': 'Запись не найдена'}), 404
        
        if record['is_excel'] and record['excel_data']:
            import io
            from openpyxl import Workbook
            
            lines = record['excel_data'].strip().split('\n')
            tasks = []
            for line in lines:
                parts = line.split('|')
                if len(parts) >= 3:
                    tasks.append({
                        'id': parts[0],
                        'title': parts[1],
                        'assignee': parts[2]
                    })
            
            wb = Workbook()
            ws = wb.active
            ws.title = "SLA Отчёт"
            
            headers = ['ID', 'Название', 'Исполнитель']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=task['id'])
                ws.cell(row=row, column=2, value=task['title'])
                ws.cell(row=row, column=3, value=task['assignee'])
            
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            
            bot_token = config.BOT_TOKEN
            chat_id = config.CHAT_ID
            
            files = {
                'document': (record['excel_filename'] or 'report.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            data = {
                'chat_id': chat_id,
                'caption': record['message_text']
            }
            
            url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
            response = requests.post(url, data=data, files=files, timeout=30)
            
            if response.status_code == 200:
                db_manager.update_notification_status(history_id, 'resent')
                return jsonify({'status': 'ok', 'message': '✅ Excel-отчёт переотправлен'})
            else:
                return jsonify({'error': f'Ошибка Telegram API: {response.status_code}'}), 500
        
        else:
            bot_token = config.BOT_TOKEN
            chat_id = config.CHAT_ID
            
            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            payload = {
                "chat_id": chat_id,
                "text": record['message_text'],
                "disable_web_page_preview": True
            }
            
            response = requests.post(url, json=payload, timeout=10)
            
            if response.status_code == 200:
                db_manager.update_notification_status(history_id, 'resent')
                return jsonify({'status': 'ok', 'message': '✅ Уведомление переотправлено'})
            else:
                return jsonify({'error': f'Ошибка Telegram API: {response.status_code}'}), 500
        
    except Exception as e:
        logger.error(f"❌ Ошибка переотправки: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/notification-history/<int:history_id>/send-pending', methods=['POST'])
def api_send_pending_notification(history_id):
    """Отправить черновик"""
    try:
        import requests
        import config
        
        record = db_manager.get_notification_by_id(history_id)
        if not record:
            return jsonify({'error': 'Запись не найдена'}), 404
        
        if record['status'] != 'pending':
            return jsonify({'error': 'Это не черновик'}), 400
        
        if record['is_excel'] and record['excel_data']:
            import io
            from openpyxl import Workbook
            
            lines = record['excel_data'].strip().split('\n')
            tasks = []
            for line in lines:
                parts = line.split('|')
                if len(parts) >= 3:
                    tasks.append({
                        'id': parts[0],
                        'title': parts[1],
                        'assignee': parts[2]
                    })
            
            wb = Workbook()
            ws = wb.active
            ws.title = "SLA Отчёт"
            
            headers = ['ID', 'Название', 'Исполнитель']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=task['id'])
                ws.cell(row=row, column=2, value=task['title'])
                ws.cell(row=row, column=3, value=task['assignee'])
            
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            
            bot_token = config.BOT_TOKEN
            chat_id = config.CHAT_ID
            
            files = {
                'document': (record['excel_filename'] or 'report.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            data = {
                'chat_id': chat_id,
                'caption': record['message_text']
            }
            
            url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
            response = requests.post(url, data=data, files=files, timeout=30)
            
            if response.status_code == 200:
                db_manager.update_notification_status(history_id, 'manual')
                return jsonify({'status': 'ok', 'message': '✅ Черновик отправлен'})
            else:
                return jsonify({'error': f'Ошибка Telegram API: {response.status_code}'}), 500
        
        else:
            bot_token = config.BOT_TOKEN
            chat_id = config.CHAT_ID
            
            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            payload = {
                "chat_id": chat_id,
                "text": record['message_text'],
                "disable_web_page_preview": True
            }
            
            response = requests.post(url, json=payload, timeout=10)
            
            if response.status_code == 200:
                db_manager.update_notification_status(history_id, 'manual')
                return jsonify({'status': 'ok', 'message': '✅ Черновик отправлен'})
            else:
                return jsonify({'error': f'Ошибка Telegram API: {response.status_code}'}), 500
        
    except Exception as e:
        logger.error(f"❌ Ошибка отправки черновика: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/notification-history/<int:history_id>', methods=['DELETE'])
def api_delete_notification(history_id):
    """Удалить запись из истории"""
    db_manager.delete_notification_history(history_id)
    return jsonify({'status': 'ok', 'message': '✅ Запись удалена'})


@app.route('/api/notification-history/clear', methods=['POST'])
def api_clear_notifications():
    """Очистить историю старше N дней (только отправленные)"""
    days = request.json.get('days', 30)
    deleted = db_manager.clear_old_notifications(days)
    return jsonify({'status': 'ok', 'message': f'✅ Удалено {deleted} записей'})


# ============ ЗАПУСК ============

if __name__ == '__main__':
    print("=" * 60)
    print("🌐 Веб-интерфейс для управления SLA ботом")
    print("📍 http://vladuha20011.fvds.ru:5000")
    print("=" * 60)
    
    app.run(host='0.0.0.0', port=5000, debug=False)
